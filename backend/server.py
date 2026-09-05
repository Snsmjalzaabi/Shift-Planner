from __future__ import annotations

import base64
import hashlib
import hmac
import html as html_escape
import io
import json
import logging
import os
import uuid
from contextlib import asynccontextmanager
from datetime import datetime, timedelta, timezone
from pathlib import Path
from typing import Any, Optional

import bcrypt
import httpx
from dotenv import load_dotenv
from fastapi import APIRouter, Depends, FastAPI, HTTPException, Request, status
from fastapi.security import HTTPAuthorizationCredentials, HTTPBearer
from jose import JWTError, jwt
from motor.motor_asyncio import AsyncIOMotorClient
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from pydantic import BaseModel, EmailStr, Field
from starlette.middleware.cors import CORSMiddleware

ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / ".env")

# ---------------------------------------------------------------------------
# Config
# ---------------------------------------------------------------------------
MONGO_URL = os.environ["MONGO_URL"]
DB_NAME = os.environ["DB_NAME"]

# SEC-002: JWT_SECRET must be provided by env; refuse to start on any default.
JWT_SECRET = os.environ.get("JWT_SECRET", "").strip()
if not JWT_SECRET or len(JWT_SECRET) < 32:
    raise RuntimeError(
        "JWT_SECRET must be set in the backend environment (>=32 chars). "
        "Generate one with: python -c \"import secrets;print(secrets.token_urlsafe(48))\""
    )
JWT_ALGORITHM = "HS256"
JWT_EXPIRES_HOURS = 24 * 7  # 7 days

# SEC-001: superuser credentials come from env only. If either is missing we
# simply skip the seed; there are no source-code defaults.
SUPERUSER_EMAIL = os.environ.get("SUPERUSER_EMAIL", "").strip()
SUPERUSER_PASSWORD = os.environ.get("SUPERUSER_PASSWORD", "").strip()

REVIEWER_EMAIL = os.environ.get("REVIEWER_EMAIL", "").strip()
REVIEWER_PASSWORD = os.environ.get("REVIEWER_PASSWORD", "").strip()

ZIINA_API_KEY = os.environ.get("ZIINA_API_KEY", "").strip()
ZIINA_API_BASE = os.environ.get("ZIINA_API_BASE", "https://api-v2.ziina.com/api").rstrip("/")
ZIINA_TEST_MODE = os.environ.get("ZIINA_TEST_MODE", "true").strip().lower() == "true"
ZIINA_PRICE_FILS = int(os.environ.get("ZIINA_PRICE_FILS", "1099"))  # 10.99 AED
ZIINA_CURRENCY = os.environ.get("ZIINA_CURRENCY", "AED").strip().upper()
ZIINA_WEBHOOK_SECRET = os.environ.get("ZIINA_WEBHOOK_SECRET", "").strip()
PLUS_PLAN_ID = "plus_monthly"
PLUS_DURATION_DAYS = 30

# CORS lockdown (P3): explicit allowlist, no wildcard.
_default_origins = "http://localhost:8081,http://localhost:19006,http://localhost:3000"
ALLOWED_ORIGINS = [
    o.strip()
    for o in os.environ.get("ALLOWED_ORIGINS", _default_origins).split(",")
    if o.strip()
]

# Private organization access. Matching accounts receive Plus automatically;
# the organization names and domains are never exposed in general app UI.
INCLUDED_ACCESS_DOMAINS = tuple(
    d.strip().lower()
    for d in os.environ.get("INCLUDED_ACCESS_DOMAINS", "").split(",")
    if d.strip()
)


def _has_included_access(email: str) -> bool:
    e = (email or "").strip().lower()
    if "@" not in e:
        return False
    domain = e.rsplit("@", 1)[1]
    return any(
        domain == d or domain.endswith("." + d)
        for d in INCLUDED_ACCESS_DOMAINS
    )

client = AsyncIOMotorClient(MONGO_URL)
db = client[DB_NAME]

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s - %(name)s - %(levelname)s - %(message)s",
)
logger = logging.getLogger("foxory")


def _client_ip(request: Request) -> str:
    """Rate-limit key that respects the ingress-proxy X-Forwarded-For header."""
    xff = request.headers.get("x-forwarded-for") or request.headers.get(
        "X-Forwarded-For"
    )
    if xff:
        return xff.split(",")[0].strip()
    return (request.client.host if request.client else "unknown") or "unknown"


# Simple in-process sliding-window rate limiter (per-process; one uvicorn
# worker per container). Fine for the MVP; swap in Redis-backed if needed.
_rate_buckets: dict[str, list[float]] = {}


def make_rate_limiter(*, name: str, per_minute: int, per_hour: int):
    async def _dep(request: Request) -> None:
        import time as _time

        now = _time.time()
        key = f"{name}:{_client_ip(request)}"
        window_hour = now - 3600.0
        hits = [t for t in _rate_buckets.get(key, []) if t >= window_hour]
        _rate_buckets[key] = hits
        recent_minute = sum(1 for t in hits if t >= now - 60.0)
        if recent_minute >= per_minute or len(hits) >= per_hour:
            raise HTTPException(
                status_code=429,
                detail="Too many requests. Please slow down and try again shortly.",
            )
        hits.append(now)

    return _dep


rate_limit_login = make_rate_limiter(name="login", per_minute=10, per_hour=60)
rate_limit_register = make_rate_limiter(name="register", per_minute=5, per_hour=30)


# ---------------------------------------------------------------------------
# Security helpers
# ---------------------------------------------------------------------------
def hash_password(password: str) -> str:
    return bcrypt.hashpw(password.encode("utf-8"), bcrypt.gensalt()).decode("utf-8")


def verify_password(password: str, hashed: str) -> bool:
    try:
        return bcrypt.checkpw(password.encode("utf-8"), hashed.encode("utf-8"))
    except Exception:
        return False


def create_access_token(sub: str, extra: dict[str, Any] | None = None) -> str:
    payload: dict[str, Any] = {
        "sub": sub,
        "iat": datetime.now(timezone.utc),
        "exp": datetime.now(timezone.utc) + timedelta(hours=JWT_EXPIRES_HOURS),
    }
    if extra:
        payload.update(extra)
    return jwt.encode(payload, JWT_SECRET, algorithm=JWT_ALGORITHM)


def decode_access_token(token: str) -> dict[str, Any]:
    return jwt.decode(token, JWT_SECRET, algorithms=[JWT_ALGORITHM])


bearer_scheme = HTTPBearer(auto_error=True)


def _as_utc(value: Any) -> Optional[datetime]:
    if isinstance(value, datetime):
        return (
            value.replace(tzinfo=timezone.utc)
            if value.tzinfo is None
            else value.astimezone(timezone.utc)
        )
    if isinstance(value, str):
        try:
            parsed = datetime.fromisoformat(value.replace("Z", "+00:00"))
            return (
                parsed.replace(tzinfo=timezone.utc)
                if parsed.tzinfo is None
                else parsed.astimezone(timezone.utc)
            )
        except ValueError:
            return None
    return None


async def _expire_paid_access_if_needed(user: dict[str, Any]) -> dict[str, Any]:
    """Downgrade an expired paid account while leaving permanent grants intact."""
    expires_at = _as_utc(user.get("plus_expires_at"))
    if (
        user.get("plan") == "plus"
        and user.get("plan_source") == "paid"
        and expires_at is not None
        and expires_at <= datetime.now(timezone.utc)
        and not user.get("is_superuser")
    ):
        now = datetime.now(timezone.utc)
        await db.users.update_one(
            {"id": user["id"]},
            {"$set": {"plan": "free", "plan_source": "free", "updated_at": now}},
        )
        user["plan"] = "free"
        user["plan_source"] = "free"
        user["updated_at"] = now
    return user


async def get_current_user(
    credentials: HTTPAuthorizationCredentials = Depends(bearer_scheme),
) -> dict[str, Any]:
    try:
        payload = decode_access_token(credentials.credentials)
    except JWTError:
        raise HTTPException(
            status_code=status.HTTP_401_UNAUTHORIZED,
            detail="Invalid or expired token",
        )
    user_id = payload.get("sub")
    if not user_id:
        raise HTTPException(status_code=status.HTTP_401_UNAUTHORIZED, detail="Invalid token payload")
    user = await db.users.find_one({"id": user_id}, {"_id": 0, "hashed_password": 0})
    if user is None:
        raise HTTPException(status_code=status.HTTP_401_UNAUTHORIZED, detail="User not found")
    return await _expire_paid_access_if_needed(user)


# ---------------------------------------------------------------------------
# Models
# ---------------------------------------------------------------------------
class RegisterRequest(BaseModel):
    email: EmailStr
    password: str = Field(min_length=6, max_length=64)
    display_name: Optional[str] = None


class LoginRequest(BaseModel):
    email: EmailStr
    password: str


class AuthResponse(BaseModel):
    access_token: str
    token_type: str = "bearer"
    user: dict[str, Any]
    registration_message: Optional[str] = None


SHIFT_TYPES = {"day", "night", "on_call", "off"}


class ShiftCreate(BaseModel):
    date: str  # YYYY-MM-DD
    type: str
    start_time: Optional[str] = None
    end_time: Optional[str] = None
    location: Optional[str] = None
    note: Optional[str] = None
    is_draft: bool = True


class ShiftUpdate(BaseModel):
    type: Optional[str] = None
    start_time: Optional[str] = None
    end_time: Optional[str] = None
    location: Optional[str] = None
    note: Optional[str] = None
    is_draft: Optional[bool] = None


class ExportRequest(BaseModel):
    month: str  # "YYYY-MM"
    include_confirmed: bool = False


# ---------------------------------------------------------------------------
# Lifespan + seed
# ---------------------------------------------------------------------------
async def seed_superuser() -> None:
    await db.users.create_index("email", unique=True)
    await db.users.create_index("id", unique=True)
    await db.shifts.create_index([("user_id", 1), ("date", 1)])

    # SEC-001: only seed when both env vars are present; never overwrite an
    # existing user's password on subsequent boots.
    if not SUPERUSER_EMAIL or not SUPERUSER_PASSWORD:
        logger.info("Superuser seed skipped (SUPERUSER_EMAIL/PASSWORD not set)")
        return

    existing = await db.users.find_one({"email": SUPERUSER_EMAIL})
    now = datetime.now(timezone.utc)
    if existing is None:
        user_doc = {
            "id": str(uuid.uuid4()),
            "email": SUPERUSER_EMAIL,
            "display_name": SUPERUSER_EMAIL.split("@")[0],
            "hashed_password": hash_password(SUPERUSER_PASSWORD),
            "is_superuser": True,
            "plan": "plus",
            "plan_source": "paid",
            "created_at": now,
            "updated_at": now,
        }
        await db.users.insert_one(user_doc)
        logger.info("Seeded superuser %s", SUPERUSER_EMAIL)
    else:
        # Only re-affirm the superuser flag; leave password + plan untouched
        # so env leakage can't silently reset the admin password.
        # Backfill plan_source for accounts seeded before the field existed.
        set_doc: dict[str, Any] = {"is_superuser": True, "updated_at": now}
        if not existing.get("plan_source"):
            set_doc["plan_source"] = "paid" if existing.get("plan") == "plus" else "free"
        await db.users.update_one({"_id": existing["_id"]}, {"$set": set_doc})
        logger.info("Superuser %s already present; flag re-affirmed", SUPERUSER_EMAIL)


async def seed_reviewer() -> None:
    """Seed an App Store / Play Store reviewer account that starts on Plus
    so reviewers can exercise every gated feature without
    paying. Never overwrites the password on subsequent boots."""
    if not REVIEWER_EMAIL or not REVIEWER_PASSWORD:
        return
    existing = await db.users.find_one({"email": REVIEWER_EMAIL})
    now = datetime.now(timezone.utc)
    if existing is None:
        await db.users.insert_one(
            {
                "id": str(uuid.uuid4()),
                "email": REVIEWER_EMAIL,
                "display_name": "App Reviewer",
                "hashed_password": hash_password(REVIEWER_PASSWORD),
                "is_superuser": False,
                "plan": "plus",
                "plan_source": "reviewer",
                "plus_activated_at": now,
                "plus_expires_at": None,
                "created_at": now,
                "updated_at": now,
            }
        )
        logger.info("Seeded reviewer account %s", REVIEWER_EMAIL)
    else:
        set_doc: dict[str, Any] = {
            "plan": "plus",
            "plan_source": "reviewer",
            "plus_expires_at": None,
            "updated_at": now,
        }
        await db.users.update_one({"_id": existing["_id"]}, {"$set": set_doc})
        logger.info("Reviewer account %s already present; plan re-affirmed", REVIEWER_EMAIL)


@asynccontextmanager
async def lifespan(app: FastAPI):
    await seed_superuser()
    await seed_reviewer()
    yield
    client.close()


app = FastAPI(title="Foxory Shift Calendar API", lifespan=lifespan)
app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origins=ALLOWED_ORIGINS,
    allow_methods=["GET", "POST", "PATCH", "DELETE", "OPTIONS"],
    allow_headers=["Authorization", "Content-Type", "X-Hmac-Signature"],
)
api_router = APIRouter(prefix="/api")


# ---------------------------------------------------------------------------
# Routes: Health & Branding
# ---------------------------------------------------------------------------
@api_router.get("/")
async def root():
    return {
        "app": "Foxory Shift Calendar",
        "created_by": "Foxory.net",
        "status": "ok",
    }


@api_router.get("/branding")
async def get_branding():
    return {
        "app_name": "Foxory Shift Calendar",
        "subtitle": "Smart shift planning for nurses and caregivers.",
        "created_by": "Foxory.net",
        "created_by_url": "https://foxory.net",
        "creator_signature": "Created by Foxory.net",
    }


# ---------------------------------------------------------------------------
# Routes: Auth
# ---------------------------------------------------------------------------
def _public_user(user: dict[str, Any]) -> dict[str, Any]:
    exp = user.get("plus_expires_at")
    return {
        "id": user["id"],
        "email": user["email"],
        "display_name": user.get("display_name"),
        "is_superuser": user.get("is_superuser", False),
        "plan": user.get("plan", "free"),
        "created_at": user["created_at"].isoformat()
        if isinstance(user["created_at"], datetime)
        else user["created_at"],
        "plus_expires_at": exp.isoformat() if isinstance(exp, datetime) else exp,
    }


async def _apply_included_access_if_needed(
    user_doc: dict[str, Any],
) -> dict[str, Any]:
    """Idempotently grant Plus to configured organization accounts."""
    if not _has_included_access(user_doc.get("email", "")):
        return user_doc
    if user_doc.get("plan") == "plus" and user_doc.get("plan_source") != "paid":
        return user_doc
    if user_doc.get("plan_source") == "paid":
        return user_doc  # they already paid — keep the paid tag
    now = datetime.now(timezone.utc)
    await db.users.update_one(
        {"id": user_doc["id"]},
        {
            "$set": {
                "plan": "plus",
                "plan_source": "included",
                "plus_activated_at": user_doc.get("plus_activated_at") or now,
                "updated_at": now,
            }
        },
    )
    user_doc["plan"] = "plus"
    user_doc["plan_source"] = "included"
    logger.info("Included organization access applied to user %s", user_doc["id"])
    return user_doc


@api_router.post(
    "/auth/register",
    response_model=AuthResponse,
    dependencies=[Depends(rate_limit_register)],
)
async def register(body: RegisterRequest):
    existing = await db.users.find_one({"email": body.email})
    if existing:
        # P3: neutral message; do not confirm/deny account existence.
        raise HTTPException(
            status_code=400,
            detail="Registration could not be completed. Please try again or sign in.",
        )
    now = datetime.now(timezone.utc)
    included_access = _has_included_access(body.email)
    plan = "plus" if included_access else "free"
    plan_source = "included" if included_access else "free"
    user_doc = {
        "id": str(uuid.uuid4()),
        "email": body.email,
        "display_name": body.display_name or body.email.split("@")[0],
        "hashed_password": hash_password(body.password),
        "is_superuser": False,
        "plan": plan,
        "plan_source": plan_source,
        "plus_activated_at": now if included_access else None,
        "created_at": now,
        "updated_at": now,
    }
    await db.users.insert_one(user_doc)
    token = create_access_token(user_doc["id"], {"email": user_doc["email"]})
    return {
        "access_token": token,
        "token_type": "bearer",
        "user": _public_user(user_doc),
        "registration_message": (
            "Thank you for supporting Foxory Shift Calendar. "
            "Please enjoy full access to the app, completely free."
            if included_access
            else None
        ),
    }


@api_router.post(
    "/auth/login",
    response_model=AuthResponse,
    dependencies=[Depends(rate_limit_login)],
)
async def login(body: LoginRequest):
    user = await db.users.find_one({"email": body.email})
    if not user or not verify_password(body.password, user["hashed_password"]):
        raise HTTPException(status_code=401, detail="Invalid email or password")
    user = await _expire_paid_access_if_needed(user)
    # Reapply private organization eligibility in case the domain policy changed.
    user = await _apply_included_access_if_needed(user)
    token = create_access_token(user["id"], {"email": user["email"]})
    return {"access_token": token, "token_type": "bearer", "user": _public_user(user)}


@api_router.get("/auth/me")
async def me(current_user: dict = Depends(get_current_user)):
    return _public_user(current_user)


class DeleteAccountRequest(BaseModel):
    password: str
    confirm: str  # must be the literal string "DELETE"


@api_router.delete("/auth/me")
async def delete_my_account(
    body: DeleteAccountRequest, current_user: dict = Depends(get_current_user)
):
    """Permanently delete the caller's account and all their data.
    Apple / Google both require this to be reachable inside the app."""
    if body.confirm != "DELETE":
        raise HTTPException(
            status_code=400,
            detail="Type DELETE (in capitals) to confirm.",
        )
    if current_user.get("is_superuser"):
        raise HTTPException(
            status_code=403,
            detail="The superuser account cannot self-delete from the app.",
        )
    # Re-verify password from the raw user record (includes hashed_password).
    raw = await db.users.find_one({"id": current_user["id"]})
    if not raw or not verify_password(body.password, raw["hashed_password"]):
        raise HTTPException(status_code=401, detail="Password is incorrect.")

    uid = current_user["id"]
    now = datetime.now(timezone.utc)
    # Best-effort cascade delete of user-owned data.
    await db.shifts.delete_many({"user_id": uid})
    await db.payments.delete_many({"user_id": uid})
    await db.email_sends.delete_many({"user_id": uid})
    await db.users.delete_one({"id": uid})
    logger.info("Account deleted: %s", current_user.get("email"))
    return {"deleted": True, "deleted_at": now.isoformat()}


def _require_plus(user: dict[str, Any]) -> None:
    if user.get("plan") != "plus" and not user.get("is_superuser"):
        raise HTTPException(
            status_code=status.HTTP_402_PAYMENT_REQUIRED,
            detail={
                "code": "plus_required",
                "message": "This feature is available on the Plus $2.99/month plan.",
            },
        )


def _current_month_str() -> str:
    d = datetime.now(timezone.utc)
    return f"{d.year:04d}-{d.month:02d}"


def _require_current_month_or_plus(user: dict[str, Any], iso_date: str) -> None:
    """Free users can only touch shifts in the current calendar month."""
    if user.get("plan") == "plus" or user.get("is_superuser"):
        return
    if not iso_date.startswith(_current_month_str() + "-"):
        raise HTTPException(
            status_code=status.HTTP_402_PAYMENT_REQUIRED,
            detail={
                "code": "plus_required",
                "message": (
                    "Multi-month planning is a Plus feature. "
                    "Upgrade to plan shifts outside the current month."
                ),
            },
        )


class PlanUpdate(BaseModel):
    plan: str = Field(pattern="^(free|plus)$")


@api_router.post("/auth/plan")
async def update_plan(
    body: PlanUpdate, current_user: dict = Depends(get_current_user)
):
    """Admin-only plan flip (used by superuser for testing).
    Regular users must go through /billing/checkout + /billing/verify."""
    if not current_user.get("is_superuser"):
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Only superusers may change plans directly. Use /billing/checkout.",
        )
    now = datetime.now(timezone.utc)
    await db.users.update_one(
        {"id": current_user["id"]},
        {"$set": {"plan": body.plan, "updated_at": now}},
    )
    fresh = await db.users.find_one(
        {"id": current_user["id"]}, {"_id": 0, "hashed_password": 0}
    )
    return _public_user(fresh)


# ---------------------------------------------------------------------------
# Billing: Ziina hosted checkout for the Plus plan
# ---------------------------------------------------------------------------
def _price_display() -> str:
    whole = ZIINA_PRICE_FILS // 100
    fract = ZIINA_PRICE_FILS % 100
    return f"{ZIINA_CURRENCY} {whole}.{fract:02d}/month"


@api_router.get("/billing/config")
async def billing_config():
    return {
        "provider": "ziina",
        "test_mode": ZIINA_TEST_MODE,
        "configured": bool(ZIINA_API_KEY),
        "currency": ZIINA_CURRENCY,
        "price_fils": ZIINA_PRICE_FILS,
        "price_display": _price_display(),
        "plans": [
            {
                "id": PLUS_PLAN_ID,
                "name": "Foxory Plus",
                "price_display": _price_display(),
                "badge_display": "Plus $2.99/month",
                "period": "month",
                "features": [
                    "XLSX export with Plan Summary + Shift Details",
                    "Attach XLSX to email exports",
                    "One-tap draft plan email from your device's mail app",
                    "Multi-month planning (any past/future month)",
                    "Priority support",
                ],
            }
        ],
    }


class CheckoutRequest(BaseModel):
    success_url: Optional[str] = None
    cancel_url: Optional[str] = None


DEFAULT_SUCCESS_URL = "foxory://billing/success"
DEFAULT_CANCEL_URL = "foxory://billing/cancel"


@api_router.post("/billing/checkout")
async def create_checkout(
    body: CheckoutRequest, current_user: dict = Depends(get_current_user)
):
    if not ZIINA_API_KEY:
        raise HTTPException(
            status_code=503,
            detail="Ziina API key not configured on the backend.",
        )
    if current_user.get("plan") == "plus":
        raise HTTPException(status_code=400, detail="Already on Plus plan.")

    success_url = body.success_url or DEFAULT_SUCCESS_URL
    cancel_url = body.cancel_url or DEFAULT_CANCEL_URL

    payload = {
        "amount": ZIINA_PRICE_FILS,
        "currency_code": ZIINA_CURRENCY,
        "success_url": success_url,
        "cancel_url": cancel_url,
        "message": f"Foxory Plus — {_price_display()}",
        "test": ZIINA_TEST_MODE,
    }
    headers = {
        "Authorization": f"Bearer {ZIINA_API_KEY}",
        "Content-Type": "application/json",
    }
    try:
        async with httpx.AsyncClient(timeout=20.0) as http:
            resp = await http.post(
                f"{ZIINA_API_BASE}/payment_intent", json=payload, headers=headers
            )
    except httpx.HTTPError as exc:
        logger.exception("Ziina checkout network error")
        raise HTTPException(status_code=502, detail=f"Ziina network error: {exc}")

    if resp.status_code >= 400:
        logger.error("Ziina checkout failed: %s %s", resp.status_code, resp.text)
        raise HTTPException(
            status_code=502,
            detail="Checkout provider is unavailable. Please try again in a moment.",
        )

    data = resp.json()
    pi_id = data.get("id")
    redirect_url = data.get("redirect_url")
    pi_status = data.get("status")
    if not pi_id or not redirect_url:
        raise HTTPException(status_code=502, detail="Malformed Ziina response.")

    now = datetime.now(timezone.utc)
    await db.payments.insert_one(
        {
            "id": pi_id,
            "user_id": current_user["id"],
            "user_email": current_user["email"],
            "amount_fils": ZIINA_PRICE_FILS,
            "currency": ZIINA_CURRENCY,
            "test_mode": ZIINA_TEST_MODE,
            "status": pi_status or "pending",
            "plan_id": PLUS_PLAN_ID,
            "success_url": success_url,
            "cancel_url": cancel_url,
            "redirect_url": redirect_url,
            "created_at": now,
            "updated_at": now,
        }
    )

    return {
        "payment_intent_id": pi_id,
        "redirect_url": redirect_url,
        "status": pi_status,
        "test_mode": ZIINA_TEST_MODE,
        "price_display": _price_display(),
    }


class VerifyRequest(BaseModel):
    payment_intent_id: str


async def _fetch_ziina_intent(intent_id: str) -> dict[str, Any]:
    headers = {"Authorization": f"Bearer {ZIINA_API_KEY}"}
    async with httpx.AsyncClient(timeout=20.0) as http:
        resp = await http.get(f"{ZIINA_API_BASE}/payment_intent/{intent_id}", headers=headers)
    if resp.status_code >= 400:
        logger.error("Ziina intent GET failed: %s %s", resp.status_code, resp.text)
        raise HTTPException(
            status_code=502,
            detail="Checkout provider is unavailable. Please try again in a moment.",
        )
    return resp.json()


async def _apply_intent_status(
    intent_id: str,
    new_status: str,
    source: str,
) -> tuple[bool, Optional[datetime], Optional[dict[str, Any]]]:
    """Idempotently reflect a Ziina payment intent status change in our DB.

    Returns (activated_this_call, plus_expires_at, user_doc) — activated is
    False if the record was already completed on a previous call."""
    payment = await db.payments.find_one({"id": intent_id}, {"_id": 0})
    if not payment:
        return False, None, None

    now = datetime.now(timezone.utc)
    already_completed = payment.get("status") == "completed"
    update_doc: dict[str, Any] = {
        "status": new_status,
        "updated_at": now,
        "last_status_source": source,
    }

    activated = False
    expires_at: Optional[datetime] = payment.get("plus_expires_at")
    fresh_user: Optional[dict[str, Any]] = None

    if new_status == "completed" and not already_completed:
        expires_at = now + timedelta(days=PLUS_DURATION_DAYS)
        update_doc.update({"completed_at": now, "plus_expires_at": expires_at})
        await db.users.update_one(
            {"id": payment["user_id"]},
            {
                "$set": {
                    "plan": "plus",
                    "plan_source": "paid",
                    "plus_expires_at": expires_at,
                    "plus_activated_at": now,
                    "updated_at": now,
                }
            },
        )
        activated = True

    await db.payments.update_one({"id": intent_id}, {"$set": update_doc})

    if activated:
        fresh_user = await db.users.find_one(
            {"id": payment["user_id"]}, {"_id": 0, "hashed_password": 0}
        )

    return activated, expires_at, fresh_user


@api_router.post("/billing/verify")
async def verify_checkout(
    body: VerifyRequest, current_user: dict = Depends(get_current_user)
):
    if not ZIINA_API_KEY:
        raise HTTPException(status_code=503, detail="Ziina API key not configured.")

    payment = await db.payments.find_one(
        {"id": body.payment_intent_id, "user_id": current_user["id"]},
        {"_id": 0},
    )
    if not payment:
        raise HTTPException(status_code=404, detail="Payment not found for this user.")

    try:
        remote = await _fetch_ziina_intent(body.payment_intent_id)
    except HTTPException:
        raise
    except Exception as exc:  # noqa: BLE001
        logger.exception("Ziina verify failed")
        raise HTTPException(status_code=502, detail=f"Ziina error: {exc}")

    new_status = remote.get("status") or payment.get("status", "pending")
    activated, expires_at, fresh_user = await _apply_intent_status(
        body.payment_intent_id, new_status, source="verify"
    )

    # If it was already completed on a prior call we still want to return the
    # latest user snapshot so the client can flip to Plus.
    if not fresh_user and new_status == "completed":
        fresh_user = await db.users.find_one(
            {"id": current_user["id"]}, {"_id": 0, "hashed_password": 0}
        )

    return {
        "payment_intent_id": body.payment_intent_id,
        "status": new_status,
        "activated": activated,
        "plus_expires_at": expires_at.isoformat() if expires_at else None,
        "user": _public_user(fresh_user) if fresh_user else None,
    }


# ---------------------------------------------------------------------------
# Ziina webhook receiver (payment_intent.status.updated)
# ---------------------------------------------------------------------------
def _verify_ziina_signature(raw_body: bytes, signature: Optional[str]) -> bool:
    # SEC-003: fail closed. Never accept unsigned webhooks — an operator who
    # rotates ZIINA_WEBHOOK_SECRET away must configure a new one.
    if not ZIINA_WEBHOOK_SECRET:
        logger.warning("Ziina webhook rejected: ZIINA_WEBHOOK_SECRET is empty")
        return False
    if not signature:
        return False
    expected = hmac.new(
        ZIINA_WEBHOOK_SECRET.encode("utf-8"),
        raw_body,
        hashlib.sha256,
    ).hexdigest()
    return hmac.compare_digest(expected, signature.strip())


@api_router.post("/billing/webhook")
async def ziina_webhook(request: Request):
    raw = await request.body()
    signature = request.headers.get("X-Hmac-Signature") or request.headers.get(
        "x-hmac-signature"
    )
    if not _verify_ziina_signature(raw, signature):
        logger.warning("Ziina webhook rejected: bad signature")
        raise HTTPException(status_code=401, detail="Invalid signature")

    try:
        event = json.loads(raw.decode("utf-8"))
    except json.JSONDecodeError:
        raise HTTPException(status_code=400, detail="Invalid JSON body")

    event_type = event.get("event_type") or event.get("event") or event.get("type")
    data = event.get("data") or event.get("payment_intent") or {}
    intent_id = data.get("id") or event.get("payment_intent_id")
    intent_status = data.get("status") or event.get("status")

    logger.info(
        "Ziina webhook: type=%s intent=%s status=%s",
        event_type,
        intent_id,
        intent_status,
    )

    # Persist every event for debugging + auditing.
    await db.webhook_events.insert_one(
        {
            "id": str(uuid.uuid4()),
            "provider": "ziina",
            "event_type": event_type,
            "intent_id": intent_id,
            "intent_status": intent_status,
            "received_at": datetime.now(timezone.utc),
            "raw": event,
        }
    )

    if not intent_id or not intent_status:
        return {"received": True, "applied": False, "reason": "missing_fields"}

    activated, _expires, _user = await _apply_intent_status(
        intent_id, intent_status, source="webhook"
    )
    return {
        "received": True,
        "intent_id": intent_id,
        "status": intent_status,
        "activated": activated,
    }


class WebhookRegisterRequest(BaseModel):
    url: str
    events: Optional[list[str]] = None


@api_router.post("/billing/webhook/register")
async def register_webhook(
    body: WebhookRegisterRequest, current_user: dict = Depends(get_current_user)
):
    """One-shot helper (superuser only) that registers this backend's
    webhook URL with Ziina using the current `ZIINA_WEBHOOK_SECRET`."""
    if not current_user.get("is_superuser"):
        raise HTTPException(status_code=403, detail="Superuser required.")
    if not ZIINA_API_KEY:
        raise HTTPException(status_code=503, detail="Ziina API key not configured.")
    if not ZIINA_WEBHOOK_SECRET:
        raise HTTPException(
            status_code=400,
            detail="Set ZIINA_WEBHOOK_SECRET in the backend .env before registering.",
        )

    payload = {
        "url": body.url,
        "secret": ZIINA_WEBHOOK_SECRET,
        "events": body.events or ["payment_intent.status.updated"],
    }
    headers = {
        "Authorization": f"Bearer {ZIINA_API_KEY}",
        "Content-Type": "application/json",
    }
    async with httpx.AsyncClient(timeout=20.0) as http:
        resp = await http.post(f"{ZIINA_API_BASE}/webhook", json=payload, headers=headers)
    if resp.status_code >= 400:
        logger.error(
            "Ziina webhook registration failed: %s %s", resp.status_code, resp.text
        )
        raise HTTPException(
            status_code=502,
            detail="Webhook provider is unavailable. Please try again in a moment.",
        )
    return {"registered": True, "ziina_response": resp.json()}


# ---------------------------------------------------------------------------
# Routes: Shifts
# ---------------------------------------------------------------------------
def _validate_shift_type(t: str) -> None:
    if t not in SHIFT_TYPES:
        raise HTTPException(status_code=400, detail=f"Invalid shift type. Must be one of {sorted(SHIFT_TYPES)}")


def _validate_date(d: str) -> None:
    try:
        datetime.strptime(d, "%Y-%m-%d")
    except ValueError:
        raise HTTPException(status_code=400, detail="date must be YYYY-MM-DD")


def _serialize_shift(doc: dict[str, Any]) -> dict[str, Any]:
    out = {k: v for k, v in doc.items() if k != "_id"}
    for k in ("created_at", "updated_at"):
        if isinstance(out.get(k), datetime):
            out[k] = out[k].isoformat()
    return out


@api_router.post("/shifts")
async def create_shift(body: ShiftCreate, current_user: dict = Depends(get_current_user)):
    _validate_shift_type(body.type)
    _validate_date(body.date)
    _require_current_month_or_plus(current_user, body.date)
    now = datetime.now(timezone.utc)
    doc = {
        "id": str(uuid.uuid4()),
        "user_id": current_user["id"],
        "date": body.date,
        "type": body.type,
        "start_time": body.start_time,
        "end_time": body.end_time,
        "location": body.location,
        "note": body.note,
        "is_draft": body.is_draft,
        "created_at": now,
        "updated_at": now,
    }
    await db.shifts.insert_one(dict(doc))
    return _serialize_shift(doc)


@api_router.get("/shifts")
async def list_shifts(
    month: Optional[str] = None,
    is_draft: Optional[bool] = None,
    current_user: dict = Depends(get_current_user),
):
    query: dict[str, Any] = {"user_id": current_user["id"]}
    if month:
        try:
            datetime.strptime(month, "%Y-%m")
        except ValueError:
            raise HTTPException(status_code=400, detail="month must be YYYY-MM")
        query["date"] = {"$regex": f"^{month}-"}
    if is_draft is not None:
        query["is_draft"] = is_draft
    cursor = db.shifts.find(query, {"_id": 0}).sort("date", 1)
    items = await cursor.to_list(length=500)
    return [_serialize_shift(it) for it in items]


@api_router.patch("/shifts/{shift_id}")
async def update_shift(
    shift_id: str, body: ShiftUpdate, current_user: dict = Depends(get_current_user)
):
    if body.type is not None:
        _validate_shift_type(body.type)
    # Load existing shift to enforce plan gating on the ORIGINAL date.
    existing = await db.shifts.find_one(
        {"id": shift_id, "user_id": current_user["id"]}, {"_id": 0}
    )
    if not existing:
        raise HTTPException(status_code=404, detail="Shift not found")
    _require_current_month_or_plus(current_user, existing["date"])
    updates = {k: v for k, v in body.dict(exclude_unset=True).items() if v is not None}
    if not updates:
        raise HTTPException(status_code=400, detail="No fields to update")
    updates["updated_at"] = datetime.now(timezone.utc)
    result = await db.shifts.find_one_and_update(
        {"id": shift_id, "user_id": current_user["id"]},
        {"$set": updates},
        return_document=True,
        projection={"_id": 0},
    )
    if not result:
        raise HTTPException(status_code=404, detail="Shift not found")
    return _serialize_shift(result)


@api_router.delete("/shifts/{shift_id}")
async def delete_shift(shift_id: str, current_user: dict = Depends(get_current_user)):
    result = await db.shifts.delete_one({"id": shift_id, "user_id": current_user["id"]})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Shift not found")
    return {"deleted": True, "id": shift_id}


@api_router.post("/shifts/{shift_id}/confirm")
async def confirm_shift(shift_id: str, current_user: dict = Depends(get_current_user)):
    result = await db.shifts.find_one_and_update(
        {"id": shift_id, "user_id": current_user["id"]},
        {"$set": {"is_draft": False, "updated_at": datetime.now(timezone.utc)}},
        return_document=True,
        projection={"_id": 0},
    )
    if not result:
        raise HTTPException(status_code=404, detail="Shift not found")
    return _serialize_shift(result)


# ---------------------------------------------------------------------------
# Routes: Export (XLSX + Email preview)
# ---------------------------------------------------------------------------
SHIFT_TYPE_LABELS = {
    "day": "Day Shift",
    "night": "Night Shift",
    "on_call": "On Call",
    "off": "Off",
}


async def _fetch_shifts_for_export(user_id: str, month: str, include_confirmed: bool) -> list[dict]:
    query: dict[str, Any] = {"user_id": user_id, "date": {"$regex": f"^{month}-"}}
    if not include_confirmed:
        query["is_draft"] = True
    cursor = db.shifts.find(query, {"_id": 0}).sort("date", 1)
    return await cursor.to_list(length=1000)


def _build_xlsx(user: dict, month: str, shifts: list[dict]) -> bytes:
    wb = Workbook()

    summary = wb.active
    summary.title = "Plan Summary"

    brand_fill = PatternFill(start_color="140C27", end_color="140C27", fill_type="solid")
    accent_font = Font(name="Calibri", size=14, bold=True, color="C084FC")
    header_font = Font(name="Calibri", size=11, bold=True, color="F8FAFC")
    body_font = Font(name="Calibri", size=11, color="F8FAFC")
    disclaimer_fill = PatternFill(start_color="1E143A", end_color="1E143A", fill_type="solid")
    disclaimer_font = Font(name="Calibri", size=10, italic=True, color="94A3B8")

    summary.column_dimensions["A"].width = 24
    summary.column_dimensions["B"].width = 60

    rows = [
        ("App Name", "Foxory Shift Calendar"),
        ("Created By", "Foxory.net"),
        ("Export Type", "Draft Planner Export"),
        ("Month", month),
        ("User", user.get("email", "")),
        ("Plan", str(user.get("plan", "free")).upper()),
        ("Generated At", datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC")),
        ("Total Shifts", str(len(shifts))),
    ]

    summary["A1"] = "Foxory Shift Calendar"
    summary["A1"].font = accent_font
    summary["A1"].fill = brand_fill
    summary["B1"] = "Draft Planner Export"
    summary["B1"].font = header_font
    summary["B1"].fill = brand_fill

    for idx, (label, value) in enumerate(rows, start=3):
        cell_a = summary.cell(row=idx, column=1, value=label)
        cell_b = summary.cell(row=idx, column=2, value=value)
        cell_a.font = header_font
        cell_a.fill = brand_fill
        cell_b.font = body_font
        cell_b.alignment = Alignment(horizontal="left")

    disclaimer_row = 3 + len(rows) + 1
    d_cell_a = summary.cell(row=disclaimer_row, column=1, value="Reminder")
    d_cell_b = summary.cell(
        row=disclaimer_row,
        column=2,
        value="This draft plan does not change the confirmed calendar.",
    )
    d_cell_a.font = header_font
    d_cell_a.fill = disclaimer_fill
    d_cell_b.font = disclaimer_font
    d_cell_b.fill = disclaimer_fill
    d_cell_b.alignment = Alignment(horizontal="left", wrap_text=True)

    signature_row = disclaimer_row + 2
    sig_cell = summary.cell(row=signature_row, column=1, value="Created by Foxory.net")
    sig_cell.font = Font(name="Calibri", size=9, italic=True, color="64748B")
    summary.merge_cells(
        start_row=signature_row,
        start_column=1,
        end_row=signature_row,
        end_column=2,
    )

    detail = wb.create_sheet("Shift Details")
    headers = ["Date", "Day", "Type", "Start", "End", "Location", "Note", "Status"]
    for col, h in enumerate(headers, start=1):
        c = detail.cell(row=1, column=col, value=h)
        c.font = header_font
        c.fill = brand_fill
        c.alignment = Alignment(horizontal="center")

    for width, letter in zip([12, 12, 14, 10, 10, 24, 40, 14], "ABCDEFGH"):
        detail.column_dimensions[letter].width = width

    for i, s in enumerate(shifts, start=2):
        try:
            d_obj = datetime.strptime(s["date"], "%Y-%m-%d").date()
            day_name = d_obj.strftime("%A")
        except Exception:
            day_name = ""
        row_vals = [
            s.get("date", ""),
            day_name,
            SHIFT_TYPE_LABELS.get(s.get("type", ""), s.get("type", "")),
            s.get("start_time") or "",
            s.get("end_time") or "",
            s.get("location") or "",
            s.get("note") or "",
            "Draft" if s.get("is_draft", True) else "Confirmed",
        ]
        for col, val in enumerate(row_vals, start=1):
            detail.cell(row=i, column=col, value=val).font = body_font

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _build_email_body(user: dict, month: str, shifts: list[dict]) -> str:
    lines = [
        f"Hi {user.get('display_name') or user.get('email', 'there')},",
        "",
        f"Here is your draft shift plan for {month}.",
        "",
        "─────────────────────────────",
    ]
    if not shifts:
        lines.append("No draft shifts have been added for this month yet.")
    else:
        for s in shifts:
            label = SHIFT_TYPE_LABELS.get(s.get("type", ""), s.get("type", ""))
            times = ""
            if s.get("start_time") and s.get("end_time"):
                times = f" · {s['start_time']}–{s['end_time']}"
            loc = f" @ {s['location']}" if s.get("location") else ""
            lines.append(f"• {s['date']}  {label}{times}{loc}")
    lines.extend(
        [
            "─────────────────────────────",
            "",
            "Reminder: This draft plan does not change the confirmed calendar.",
            "",
            "—",
            "This draft shift plan was created using Foxory Shift Calendar — created by Foxory.net.",
        ]
    )
    return "\n".join(lines)


def _build_email_html(user: dict, month: str, shifts: list[dict]) -> str:
    def _e(v: Any) -> str:
        return html_escape.escape(str(v or ""))

    rows_html = ""
    if not shifts:
        rows_html = (
            '<tr><td colspan="4" style="padding:16px;color:#94A3B8;'
            'font-style:italic;text-align:center;">'
            "No draft shifts have been added for this month yet.</td></tr>"
        )
    else:
        for s in shifts:
            label = SHIFT_TYPE_LABELS.get(s.get("type", ""), s.get("type", ""))
            time_str = ""
            if s.get("start_time") or s.get("end_time"):
                time_str = f"{_e(s.get('start_time') or '—')} → {_e(s.get('end_time') or '—')}"
            rows_html += (
                "<tr>"
                f'<td style="padding:10px 12px;border-bottom:1px solid #221641;color:#F8FAFC;font-weight:600;">{_e(s.get("date"))}</td>'
                f'<td style="padding:10px 12px;border-bottom:1px solid #221641;color:#C084FC;">{_e(label)}</td>'
                f'<td style="padding:10px 12px;border-bottom:1px solid #221641;color:#94A3B8;font-family:monospace;">{time_str}</td>'
                f'<td style="padding:10px 12px;border-bottom:1px solid #221641;color:#94A3B8;">{_e(s.get("location") or "")}</td>'
                "</tr>"
            )

    greet = _e(user.get("display_name") or user.get("email", "there"))

    return f"""<!doctype html>
<html>
  <body style="margin:0;padding:0;background:#090514;font-family:-apple-system,BlinkMacSystemFont,'Segoe UI',Roboto,sans-serif;">
    <table role="presentation" width="100%" cellspacing="0" cellpadding="0" style="background:#090514;padding:24px 0;">
      <tr>
        <td align="center">
          <table role="presentation" width="600" cellspacing="0" cellpadding="0" style="max-width:600px;width:100%;background:#140C27;border:1px solid #2D1B4E;border-radius:16px;overflow:hidden;">
            <tr>
              <td style="padding:24px 28px 8px;">
                <div style="color:#C084FC;font-size:11px;letter-spacing:2px;text-transform:uppercase;font-weight:700;">Foxory Shift Calendar</div>
                <div style="color:#F8FAFC;font-size:22px;font-weight:800;margin-top:4px;">Draft plan for {_e(month)}</div>
              </td>
            </tr>
            <tr>
              <td style="padding:8px 28px 0;color:#94A3B8;font-size:14px;line-height:22px;">
                Hi {greet},<br/>
                Here is your draft shift plan.
              </td>
            </tr>
            <tr>
              <td style="padding:16px 20px;">
                <table role="presentation" width="100%" cellspacing="0" cellpadding="0" style="background:#0C0620;border:1px solid #221641;border-radius:12px;overflow:hidden;">
                  <thead>
                    <tr>
                      <th align="left" style="padding:10px 12px;color:#64748B;font-size:10px;letter-spacing:1px;text-transform:uppercase;font-weight:700;background:#0F0827;">Date</th>
                      <th align="left" style="padding:10px 12px;color:#64748B;font-size:10px;letter-spacing:1px;text-transform:uppercase;font-weight:700;background:#0F0827;">Type</th>
                      <th align="left" style="padding:10px 12px;color:#64748B;font-size:10px;letter-spacing:1px;text-transform:uppercase;font-weight:700;background:#0F0827;">Time</th>
                      <th align="left" style="padding:10px 12px;color:#64748B;font-size:10px;letter-spacing:1px;text-transform:uppercase;font-weight:700;background:#0F0827;">Location</th>
                    </tr>
                  </thead>
                  <tbody>
                    {rows_html}
                  </tbody>
                </table>
              </td>
            </tr>
            <tr>
              <td style="padding:4px 28px 16px;color:#94A3B8;font-size:12px;font-style:italic;">
                Reminder: This draft plan does not change the confirmed calendar.
              </td>
            </tr>
            <tr>
              <td style="padding:0 28px 24px;">
                <div style="border-top:1px solid #2D1B4E;padding-top:16px;color:#64748B;font-size:11px;">
                  This draft shift plan was created using
                  <span style="color:#94A3B8;font-weight:600;">Foxory Shift Calendar</span>
                  — created by
                  <a href="https://foxory.net" style="color:#C084FC;text-decoration:underline;">Foxory.net</a>.
                </div>
              </td>
            </tr>
          </table>
        </td>
      </tr>
    </table>
  </body>
</html>"""


@api_router.post("/export/xlsx")
async def export_xlsx(body: ExportRequest, current_user: dict = Depends(get_current_user)):
    _require_plus(current_user)
    try:
        datetime.strptime(body.month, "%Y-%m")
    except ValueError:
        raise HTTPException(status_code=400, detail="month must be YYYY-MM")
    shifts = await _fetch_shifts_for_export(current_user["id"], body.month, body.include_confirmed)
    data = _build_xlsx(current_user, body.month, shifts)
    b64 = base64.b64encode(data).decode("ascii")
    filename = f"foxory-shift-plan-{body.month}.xlsx"
    return {
        "filename": filename,
        "content_type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        "base64": b64,
        "size_bytes": len(data),
        "shift_count": len(shifts),
        "created_by": "Foxory.net",
    }


@api_router.post("/export/email")
async def export_email(body: ExportRequest, current_user: dict = Depends(get_current_user)):
    """Render a draft-plan email (subject / plain-text body / HTML body).

    Delivery is handled on the client via the native mail composer
    (`expo-mail-composer`) which opens the user's own mail app pre-filled.
    This endpoint no longer relays mail; it is a pure rendering endpoint.
    """
    try:
        datetime.strptime(body.month, "%Y-%m")
    except ValueError:
        raise HTTPException(status_code=400, detail="month must be YYYY-MM")
    shifts = await _fetch_shifts_for_export(current_user["id"], body.month, body.include_confirmed)
    subject = f"Foxory Shift Calendar — Draft Plan for {body.month}"
    body_text = _build_email_body(current_user, body.month, shifts)
    body_html = _build_email_html(current_user, body.month, shifts)
    return {
        "to": current_user["email"],
        "subject": subject,
        "body": body_text,
        "html": body_html,
        "shift_count": len(shifts),
        "signature": "Created by Foxory.net",
    }


app.include_router(api_router)
