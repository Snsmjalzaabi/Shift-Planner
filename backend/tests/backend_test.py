"""Foxory Shift Calendar - backend integration tests."""
import base64
import io
import os
import uuid
from datetime import datetime

import pytest
import requests
from openpyxl import load_workbook

BASE_URL = os.environ["EXPO_PUBLIC_BACKEND_URL"].rstrip("/") if os.environ.get(
    "EXPO_PUBLIC_BACKEND_URL"
) else "https://nurse-planner-5.preview.emergentagent.com"
API = f"{BASE_URL}/api"

SUPER_EMAIL = "Sultan942002@yahoo.com"
SUPER_PASS = "S.nsmjalzaabi1"
MONTH = "2026-03"

session = requests.Session()
session.headers.update({"Content-Type": "application/json"})


@pytest.fixture(scope="session")
def super_token():
    r = session.post(f"{API}/auth/login", json={"email": SUPER_EMAIL, "password": SUPER_PASS}, timeout=30)
    assert r.status_code == 200, f"login failed: {r.status_code} {r.text}"
    data = r.json()
    assert data["user"]["plan"] == "plus"
    assert data["user"]["is_superuser"] is True
    return data["access_token"]


# ---------- branding ----------
def test_branding():
    r = session.get(f"{API}/branding", timeout=15)
    assert r.status_code == 200
    d = r.json()
    assert d["app_name"] == "Foxory Shift Calendar"
    assert d["subtitle"] == "Smart shift planning for nurses and caregivers."
    assert d["created_by"] == "Foxory.net"
    assert d["creator_signature"] == "Created by Foxory.net"


# ---------- auth ----------
def test_login_superuser():
    r = session.post(f"{API}/auth/login", json={"email": SUPER_EMAIL, "password": SUPER_PASS}, timeout=15)
    assert r.status_code == 200
    d = r.json()
    assert "access_token" in d and d["access_token"]
    assert d["user"]["email"] == SUPER_EMAIL
    assert d["user"]["plan"] == "plus"
    assert d["user"]["is_superuser"] is True


def test_me(super_token):
    r = session.get(f"{API}/auth/me", headers={"Authorization": f"Bearer {super_token}"}, timeout=15)
    assert r.status_code == 200
    d = r.json()
    assert d["email"] == SUPER_EMAIL
    assert d["plan"] == "plus"
    assert d["is_superuser"] is True


def test_register_new_user():
    email = f"TEST_{uuid.uuid4().hex[:8]}@example.com"
    r = session.post(
        f"{API}/auth/register",
        json={"email": email, "password": "testpass123", "display_name": "TEST User"},
        timeout=15,
    )
    assert r.status_code == 200, r.text
    d = r.json()
    assert d["user"]["plan"] == "free"
    assert d["user"]["is_superuser"] is False
    assert d["access_token"]


def test_login_invalid():
    r = session.post(f"{API}/auth/login", json={"email": SUPER_EMAIL, "password": "wrong"}, timeout=15)
    assert r.status_code == 401


# ---------- shifts CRUD ----------
def _auth(t):
    return {"Authorization": f"Bearer {t}"}


def test_shifts_full_crud(super_token):
    hdr = _auth(super_token)
    date = f"{MONTH}-15"

    # Create
    r = session.post(
        f"{API}/shifts",
        json={"date": date, "type": "day", "start_time": "07:00", "end_time": "19:00", "location": "Ward 3", "note": "TEST", "is_draft": True},
        headers=hdr, timeout=15,
    )
    assert r.status_code == 200, r.text
    shift = r.json()
    sid = shift["id"]
    assert shift["is_draft"] is True
    assert shift["type"] == "day"

    # List by month
    r = session.get(f"{API}/shifts?month={MONTH}", headers=hdr, timeout=15)
    assert r.status_code == 200
    items = r.json()
    assert any(s["id"] == sid for s in items)

    # Patch
    r = session.patch(f"{API}/shifts/{sid}", json={"location": "ICU"}, headers=hdr, timeout=15)
    assert r.status_code == 200
    assert r.json()["location"] == "ICU"

    # Confirm
    r = session.post(f"{API}/shifts/{sid}/confirm", headers=hdr, timeout=15)
    assert r.status_code == 200
    assert r.json()["is_draft"] is False

    # Delete
    r = session.delete(f"{API}/shifts/{sid}", headers=hdr, timeout=15)
    assert r.status_code == 200
    r = session.delete(f"{API}/shifts/{sid}", headers=hdr, timeout=15)
    assert r.status_code == 404


def test_shifts_invalid_type(super_token):
    r = session.post(
        f"{API}/shifts",
        json={"date": "2026-03-01", "type": "invalid"},
        headers=_auth(super_token), timeout=15,
    )
    assert r.status_code == 400


# ---------- export xlsx ----------
def test_export_xlsx(super_token):
    hdr = _auth(super_token)
    # ensure at least one shift
    date = f"{MONTH}-20"
    r = session.post(
        f"{API}/shifts",
        json={"date": date, "type": "night", "start_time": "19:00", "end_time": "07:00", "location": "TEST-LOC", "is_draft": True},
        headers=hdr, timeout=15,
    )
    assert r.status_code == 200
    sid = r.json()["id"]

    try:
        r = session.post(
            f"{API}/export/xlsx",
            json={"month": MONTH, "include_confirmed": False},
            headers=hdr, timeout=30,
        )
        assert r.status_code == 200, r.text
        d = r.json()
        assert d["filename"] == f"foxory-shift-plan-{MONTH}.xlsx"
        assert d["created_by"] == "Foxory.net"
        assert d["shift_count"] >= 1
        assert d["base64"]

        # decode and inspect
        raw = base64.b64decode(d["base64"])
        assert len(raw) > 100
        wb = load_workbook(io.BytesIO(raw))
        assert "Plan Summary" in wb.sheetnames
        ws = wb["Plan Summary"]
        cells = []
        for row in ws.iter_rows(values_only=True):
            cells.extend([c for c in row if c is not None])
        text_blob = "\n".join(str(c) for c in cells)
        assert "Foxory Shift Calendar" in text_blob
        assert "Foxory.net" in text_blob
        assert "Draft Planner Export" in text_blob
        assert "App Name" in text_blob
        assert "Created By" in text_blob
        assert "Export Type" in text_blob
        assert "This draft plan does not change the confirmed calendar." in text_blob
    finally:
        session.delete(f"{API}/shifts/{sid}", headers=hdr, timeout=15)


# ---------- export email ----------
def test_export_email(super_token):
    hdr = _auth(super_token)
    r = session.post(
        f"{API}/export/email",
        json={"month": MONTH, "include_confirmed": True},
        headers=hdr, timeout=15,
    )
    assert r.status_code == 200, r.text
    d = r.json()
    assert d["to"] == SUPER_EMAIL
    assert MONTH in d["subject"]
    assert "Foxory Shift Calendar" in d["subject"]
    body = d["body"]
    assert body.rstrip().endswith(
        "This draft shift plan was created using Foxory Shift Calendar — created by Foxory.net."
    )
    assert "Reminder: This draft plan does not change the confirmed calendar." in body
    assert d["signature"] == "Created by Foxory.net"


def test_export_invalid_month(super_token):
    r = session.post(
        f"{API}/export/xlsx",
        json={"month": "2026-13"},
        headers=_auth(super_token), timeout=15,
    )
    assert r.status_code == 400


def test_unauthorized_shifts():
    r = session.get(f"{API}/shifts", timeout=15)
    assert r.status_code in (401, 403)
